from io import StringIO
from pathlib import Path
import time
import pandas as pd
import requests
from openpyxl import load_workbook

from config.series import TREASURY_PAGES, TARGET_COLUMNS

TIMEOUT = 30
MAX_RETRIES = 3


WORKBOOK_PATH = Path(__file__).resolve().parent.parent / "data" / "master" / "US Yields.xlsx"


def get_last_date_from_sheet(workbook_path: Path, sheet_name: str = "Data") -> pd.Timestamp:
    """
    Read the latest valid date from the workbook's Data sheet.
    Assumes the first column contains dates.
    """
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[sheet_name]

    dates = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None:
            try:
                dates.append(pd.to_datetime(row[0]))
            except Exception:
                pass

    if not dates:
        raise ValueError(f"No valid dates found in sheet '{sheet_name}'")

    return max(dates)


def fetch_treasury_table(url: str) -> pd.DataFrame:
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = requests.get(url, timeout=TIMEOUT)
            response.raise_for_status()
            tables = pd.read_html(StringIO(response.text), flavor="html5lib")
            break
        except Exception as e:
            if attempt == MAX_RETRIES:
                raise RuntimeError(f"Failed to fetch {url} after {MAX_RETRIES} attempts: {e}") from e
            print(f"Attempt {attempt} failed: {e}. Retrying...")
            time.sleep(2 ** attempt)

    if not tables:
        raise ValueError(f"No table found at {url}")

    df = tables[0].copy()
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df = df.dropna(subset=["Date"])
    df = df.sort_values("Date").reset_index(drop=True)

    return df
def keep_target_columns(df: pd.DataFrame, dataset_name: str) -> pd.DataFrame:
    """
    Keep only the columns we want, in the exact order we want.
    """
    wanted = TARGET_COLUMNS[dataset_name]

    missing = [col for col in wanted if col not in df.columns]
    if missing:
        print(f"Warning: missing columns in {dataset_name}: {missing}")

    existing = [col for col in wanted if col in df.columns]
    return df[existing].copy()

def convert_to_weekly(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert daily Treasury data to weekly Friday observations.
    """
    weekly = (
        df.set_index("Date")
          .resample("W-FRI")
          .last()
          .reset_index()
    )
    return weekly


def filter_new_rows(df_weekly: pd.DataFrame, last_existing_date: pd.Timestamp) -> pd.DataFrame:
    """
    Keep only rows newer than the latest date already in Excel.
    """
    return df_weekly[df_weekly["Date"] > last_existing_date].copy()


NOMINAL_SHEET = "treasury_rates_2024-2026"
TIPS_SHEET = "TIPS 2020_2026"


def save_output(df: pd.DataFrame, name: str) -> None:
    out_dir = Path(f"data/weekly/{name}")
    out_dir.mkdir(parents=True, exist_ok=True)
    df.to_csv(out_dir / f"{name}_new_rows.csv", index=False)


def get_last_row(ws) -> int:
    for row in range(ws.max_row, 0, -1):
        if any(cell.value is not None for cell in ws[row]):
            return row
    return 1


def append_to_sheet(ws, df: pd.DataFrame) -> None:
    start_row = get_last_row(ws) + 1
    for i, row in enumerate(df.itertuples(index=False), start=start_row):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)


def update_excel(nominal_df: pd.DataFrame, tips_df: pd.DataFrame) -> None:
    wb = load_workbook(WORKBOOK_PATH)

    print(f"Nominal rows to add: {len(nominal_df)}")
    print(f"TIPS rows to add: {len(tips_df)}")

    if not nominal_df.empty:
        append_to_sheet(wb[NOMINAL_SHEET], nominal_df)
        print("Nominal data appended.")

    if not tips_df.empty:
        append_to_sheet(wb[TIPS_SHEET], tips_df)
        print("TIPS data appended.")

    wb.save(WORKBOOK_PATH)
    print("Workbook updated successfully.")


def main():
    last_date = get_last_date_from_sheet(WORKBOOK_PATH, sheet_name="Data")
    print(f"Latest date already in workbook: {last_date.date()}")

    print("Downloading current-year nominal Treasury data...")
    nominal_daily = fetch_treasury_table(TREASURY_PAGES["nominal"])
    nominal_daily = keep_target_columns(nominal_daily, "nominal")
    nominal_weekly = convert_to_weekly(nominal_daily)
    nominal_new = filter_new_rows(nominal_weekly, last_date)
    save_output(nominal_new, "nominal_yield_curve")

    print("Downloading current-year TIPS Treasury data...")
    tips_daily = fetch_treasury_table(TREASURY_PAGES["tips"])
    tips_daily = keep_target_columns(tips_daily, "tips")
    tips_weekly = convert_to_weekly(tips_daily)
    tips_new = filter_new_rows(tips_weekly, last_date)
    save_output(tips_new, "tips_yield_curve")

    print(f"Nominal new rows: {len(nominal_new)}")
    print(f"TIPS new rows: {len(tips_new)}")

    update_excel(nominal_new, tips_new)


if __name__ == "__main__":
    main()
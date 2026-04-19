from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

WORKBOOK_PATH = Path(__file__).resolve().parent.parent / "data" / "master" / "US Yields.xlsx"

NOMINAL_SHEET = "treasury_rates_2024-2026"
TIPS_SHEET = "TIPS 2020_2026"

def get_last_row(ws):
    """
    Find last non-empty row
    """
    for row in range(ws.max_row, 0, -1):
        if any(cell.value is not None for cell in ws[row]):
            return row
    return 1


def append_dataframe(ws, df: pd.DataFrame):
    """
    Append dataframe to worksheet
    """
    start_row = get_last_row(ws) + 1

    for i, row in enumerate(df.itertuples(index=False), start=start_row):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)


def load_new_data(path):
    if not path.exists():
        print(f"No file: {path}")
        return pd.DataFrame()

    df = pd.read_csv(path)
    df["Date"] = pd.to_datetime(df["Date"])
    return df


def main():
    wb = load_workbook(WORKBOOK_PATH)

    # Load new data
    nominal_path = Path("data/weekly/nominal_yield_curve/nominal_yield_curve_new_rows.csv")
    tips_path = Path("data/weekly/tips_yield_curve/tips_yield_curve_new_rows.csv")

    nominal_df = load_new_data(nominal_path)
    tips_df = load_new_data(tips_path)

    print(f"Nominal rows to add: {len(nominal_df)}")
    print(f"TIPS rows to add: {len(tips_df)}")

    # Append nominal
    if not nominal_df.empty:
        ws_nominal = wb[NOMINAL_SHEET]
        append_dataframe(ws_nominal, nominal_df)
        print("Nominal data appended")

    # Append TIPS
    if not tips_df.empty:
        ws_tips = wb[TIPS_SHEET]
        append_dataframe(ws_tips, tips_df)
        print("TIPS data appended")

    wb.save(WORKBOOK_PATH)

    print("Workbook updated successfully.")


if __name__ == "__main__":
    main()